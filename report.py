from tkinter import*
import sqlite3 as sql
import openpyxl as xl
from datetime import datetime, timedelta
from tkinter import messagebox as mb

conn = sql.connect('Hotel.db')
#cur=conn.cursor()

rootRep = Tk()
rootRep.title('Отчёты')


def reportToExcelFirst():
    wb = xl.Workbook()
    wb.remove(wb['Sheet'])
    
    sheet = wb.create_sheet("Список свободных комнат")
    cursor = conn.execute("Select RoomID, Type, Price_day, Employment, Booking From Rooms Where Employment='Свободно'")
    curs = conn.execute("select count(RoomID) from Rooms Where Employment='Свободно'")
    sheet.cell(row=1, column=2).value="Список свободных комнат"
    sheet.cell(row=2, column=1).value="Номер комнаты"
    sheet.cell(row=2, column=2).value="Тип"
    sheet.cell(row=2, column=3).value="Цена в сутки"
    sheet.cell(row=2, column=4).value="Статус"
    sheet.cell(row=2, column=5).value="Бронь"
    sheet.cell(row=6, column=8).value="Всего свободных комнат"
    sheet.cell(row=6, column=9).value=curs.fetchone()[0]
    sheet.cell(row=4, column=8).value="Информация создана "+str(datetime.today())
    i=3
    for roq in cursor:
        sheet.cell(row=i, column=1).value=roq[0]
        sheet.cell(row=i, column=2).value=roq[1]
        sheet.cell(row=i, column=3).value=roq[2]
        sheet.cell(row=i, column=4).value=roq[3]
        sheet.cell(row=i, column=5).value=roq[4]
        i+=1
    conn.commit()
    
    sheet1 = wb.create_sheet("Список занятых комнат")
    cursor1 = conn.execute("Select RoomID, Type, Price_day, Employment, Booking From Rooms Where Employment='Занято'")
    curs = conn.execute("select count(RoomID) from Rooms Where Employment='Занято'")
    sheet1.cell(row=1, column=2).value="Список занятых комнат"
    sheet1.cell(row=2, column=1).value="Номер комнаты"
    sheet1.cell(row=2, column=2).value="Тип"
    sheet1.cell(row=2, column=3).value="Цена в сутки"
    sheet1.cell(row=2, column=4).value="Статус"
    sheet1.cell(row=2, column=5).value="Бронь"
    sheet1.cell(row=6, column=8).value="Всего занятых комнат"
    sheet1.cell(row=6, column=9).value=curs.fetchone()[0]
    sheet1.cell(row=4, column=8).value="Информация создана "+str(datetime.today())
    i1=3
    for roq1 in cursor1:
        sheet1.cell(row=i1, column=1).value=roq1[0]
        sheet1.cell(row=i1, column=2).value=roq1[1]
        sheet1.cell(row=i1, column=3).value=roq1[2]
        sheet1.cell(row=i1, column=4).value=roq1[3]
        sheet1.cell(row=i1, column=5).value=roq1[4]
        i1+=1
    conn.commit()
    d1 = datetime.now() - timedelta(days=30)
    dnow1 = datetime.now()
    sheet2 = wb.create_sheet("Деньги с каждого номера")
    cursor2 = conn.execute("Select RoomNum, Sum(Payment), Status FROM Work Group by RoomNum")
    curs = conn.execute("select Sum(Payment) from Work")
    sheet2.cell(row=1, column=2).value="Деньги с каждого номера"
    sheet2.cell(row=2, column=1).value="Номер комнаты"
    sheet2.cell(row=2, column=2).value="Сумма"
    sheet2.cell(row=2, column=3).value="Цена в сутки"
    sheet2.cell(row=6, column=6).value="Сумма со всех номеров"
    sheet2.cell(row=6, column=7).value=curs.fetchone()[0]
    sheet2.cell(row=3, column=5).value="Информация создана "+str(datetime.today())
    i2=3
    for roq2 in cursor2:
        sheet2.cell(row=i2, column=1).value=roq2[0]
        sheet2.cell(row=i2, column=2).value=roq2[1]
        sheet2.cell(row=i2, column=3).value=roq2[2]
        i2+=1
    conn.commit()
    d = datetime.now() - timedelta(days=7)
    dnow = datetime.now()
    sheet3 = wb.create_sheet("Общая выручка за неделю")
    cursor3 = conn.execute("Select Sum(Payment) FROM Work Where Status ='Оплачено' and DateStart between '"+d.strftime("%d-%m-%Y")+"' and '"+dnow.strftime("%d-%m-%Y")+"'")
    sheet3.cell(row=1, column=2).value="Общая выручка за неделю"
    sheet3.cell(row=2, column=1).value="Общая сумма"
    sheet3.cell(row=3, column=2).value="руб."
    i3=3
    for roq3 in cursor3:
        sheet3.cell(row=i3, column=1).value=roq3[0]
        i3+=1
    wb.save('Списки комнат.xlsx')
    conn.commit()
    mb.showinfo("Успешно", "Отчёт был создан и сохранён в корневой папке")


def reportToExcelSecond():
    wb=xl.Workbook()
    wb.remove(wb["Sheet"])

    sheet=wb.create_sheet("Рейтинг клиентов")
    curs = conn.execute("Select ClientNum, Surname || ' ' || Name, count(ClientNum) FROM Work INNER JOIN Clients ON ClientID = ClientNum Group by ClientNum Order by count(ClientNum) DESC")
    cu = conn.execute("Select count(Surname || ' ' || Name) From Work Inner join Clients ON ClientID = ClientNum")
    sheet.cell(row=1, column=2).value="Рейтинг клиентов"
    sheet.cell(row=2, column=1).value="Номер клиента"
    sheet.cell(row=2, column=2).value="ФИО клиента"
    sheet.cell(row=2, column=3).value="Кол-во оформлений"
    sheet.cell(row=5, column=5).value="Общее количество оформлений"
    sheet.cell(row=5, column=6).value=cu.fetchone()[0]
    sheet.cell(row=3, column=5).value="Информация создана "+str(datetime.today())
    i=2
    for roq in curs:
        sheet.cell(row=i, column=1).value=roq[0]
        sheet.cell(row=i, column=2).value=roq[1]
        sheet.cell(row=i, column=3).value=roq[2]
        i+=1
    wb.save('Данные клиентов.xlsx')
    conn.commit()
    mb.showinfo("Успешно", "Отчёт был составлен и сохранён в коневой папке")

b1 = Button(rootRep, text = "Рейтинг клиентов", width=35, command=reportToExcelSecond).pack(side='top',padx=5, pady=10)

b2 = Button(rootRep, text = "Списки комнат", width=35, command=reportToExcelFirst).pack(side='top',padx=5, pady=10)
